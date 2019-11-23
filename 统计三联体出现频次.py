import os
import openpyxl
from openpyxl.writer.excel import ExcelWriter

# filePath=r'C:\Users\lenovo\Desktop'                           #r'C:\Users\user\Desktop\找三联体bindingsite的规律\无二级结构的rna结构'
filePath = r'./'  # r'C:\Users\user\Desktop\找三联体bindingsite的规律\无二级结构的rna结构'
workBook = openpyxl.load_workbook(os.path.join(filePath, r'protein实验位点标记数据11.20.xlsx'), read_only=False)
sheetList = workBook.get_sheet_names()

expRes = {}
preRes = {}
# 颜色预设
red = openpyxl.styles.Font(name='Arial', size=11, bold=False, italic=False, vertAlign=None, underline='none',
                           strike=False, color='00FF0000')
black = openpyxl.styles.Font(name='Arial', size=11, bold=False, italic=False, vertAlign=None, underline='none',
                             strike=False, color='00000000')
green = openpyxl.styles.Font(name='Arial', size=11, bold=False, italic=False, vertAlign=None, underline='none',
                             strike=False, color='0000FF00')

columnC = 3  # expriment bindingsite
columnI = 9  # triplet
columnJ = 10  # prediction
columnM = 13  # pdb seq


def rnaNameAppend(expPattern, prePattern, dict, rnaName, tripletName):
    index = 0
    for j in range(prePattern):
        index += dict[tripletName][expPattern][j]
    dict[tripletName][expPattern].insert(index + 8, rnaName)


def printResultXls(row, rowIndex, column, pattern, triplet, color):
    """
    根据规定的颜色输出三联体单元格
    """
    ws.cell(row + rowIndex, column + 4).value = triplet[0]
    ws.cell(row + rowIndex, column + 5).value = triplet[1]
    ws.cell(row + rowIndex, column + 6).value = triplet[2]

    if pattern >> 2 & 1 == 1:
        ws.cell(row + rowIndex, column + 4).font = color
    else:
        pass

    if pattern >> 1 & 1 == 1:
        ws.cell(row + rowIndex, column + 5).font = color
    else:
        pass

    if pattern & 1 == 1:
        ws.cell(row + rowIndex, column + 6).font = color
    else:
        pass


for patternNo in sheetList:
    print(patternNo)

    curSheet = workBook[patternNo]
    tripletPattern = 0
    curRow = 3
    preRow = 3
    # 向res 中添加rna名称 和 三联体
    while curRow < curSheet.max_row:
        if curSheet.cell(curRow, columnI).value == 1:
            if curRow - preRow == 3:
                tripletName = curSheet.cell(curRow - 2, columnM).value + curSheet.cell(curRow - 1, columnM).value + \
                              curSheet.cell(curRow, columnM).value
                # 统计实验模式
                tripletPattern = int(curSheet.cell(curRow - 2, columnC).value) * 4 + \
                                 int(curSheet.cell(curRow - 1, columnC).value) * 2 + \
                                 int(curSheet.cell(curRow, columnC).value)
                expRes.setdefault(tripletName, [[], [], [], [], [], [], [], []])[tripletPattern].append(patternNo)
                # 预测位点统计
                prePattern = int(curSheet.cell(curRow - 2, columnJ).value) * 4 + \
                             int(curSheet.cell(curRow - 1, columnJ).value) * 2 + \
                             int(curSheet.cell(curRow, columnJ).value)

                preRes.setdefault(tripletName, [[0 for i in range(8)], [0 for i in range(8)], [0 for i in range(8)],
                                                [0 for i in range(8)], [0 for i in range(8)], [0 for i in range(8)],
                                                [0 for i in range(8)], [0 for i in range(8)]])[tripletPattern][
                    prePattern] += 1
                # preRes[tripletName][tripletPattern].append(i)
                rnaNameAppend(tripletPattern, prePattern, preRes, patternNo, tripletName)
                preRow += 1
        else:
            preRow = curRow

        curRow += 1
        tripletPattern = 0

# 输出excel表格
# for i in preRes:
#    print(preRes)
resWorkBook = openpyxl.Workbook()
# ew=ExcelWriter(workbook=resWorkBook)
ws = resWorkBook.create_sheet('protein')

baseRow = 2
baseColumn = 1
for triplet in expRes:
    ws.cell(baseRow, baseColumn).value = triplet
    totalSum = 0
    preSum = 0
    for patternNo in range(8):
        expRnaConnSet = ''
        totalSum += len(expRes[triplet][patternNo])
        expSum = len(expRes[triplet][patternNo])
        rowOffset = 0
        if expSum != 0:
            for rnaIdx in range(expSum):
                if rnaIdx == 0:
                    expRnaConnSet += expRes[triplet][patternNo][rnaIdx]
                else:
                    expRnaConnSet += '\\' + expRes[triplet][patternNo][rnaIdx]
            # 写rna名称
            ws.cell(baseRow + rowOffset, baseColumn + 3).value = expRnaConnSet

            # 写 三联体实验位点的 不同模式以及出现次数
            printResultXls(baseRow, rowOffset, baseColumn, patternNo, triplet, red)
            ws.cell(baseRow + rowOffset + 1, baseColumn + 4).value = expSum

            # ws.merge_cells(start_row=row+rowIndex+1,start_column=column+4,end_row=row+rowIndex+1,end_column=column+6)
            # 写 三联体预测位点的不同模式、对应的rna名称以及出现次数
            colOffset = 3
            for rnaSumIdx in range(8):
                if preRes[triplet][patternNo][rnaSumIdx] != 0:
                    # 写 预测位点模式
                    printResultXls(baseRow, rowOffset, baseColumn + 4 + colOffset, patternNo, triplet, green)
                    # 写 出现次数
                    ws.cell(baseRow + rowOffset + 1, baseColumn + 4 + colOffset).value = preRes[triplet][patternNo][rnaSumIdx]
                    # 预测位点等于实验位点时不输出
                    if preRes[triplet][patternNo][rnaSumIdx] == expSum:
                        continue
                    else:
                        # 写rna名称
                        ws.cell(baseRow + rowOffset + 1, baseColumn + 5 + colOffset).value = '\\'.join(
                            preRes[triplet][patternNo][rnaSumIdx + 8:rnaSumIdx + 8 + preRes[triplet][patternNo][rnaSumIdx]])
                        colOffset += 3

            # 写完一个pattern，将行偏移加2，准备写下一个pattern
            rowOffset += 2
            # 行基址也加2，把下一个rna往下顶
            baseRow += 2

    # 写三联体出现次数
    ws.cell(baseRow, baseColumn + 1).value = totalSum
    # 写带实验位点的三联体出现次数
    ws.cell(baseRow, baseColumn + 2).value = totalSum - len(expRes[triplet][0])

    # baseColumn += 1
    baseRow += 2

print('hi')
resWorkBook.save(filePath + 'test.xlsx')

print('hello')
